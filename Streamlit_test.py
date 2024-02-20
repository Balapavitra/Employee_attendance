import streamlit as st
import face_recognition
import cv2
import os
from datetime import datetime
import openpyxl

# Function to load employee faces
def load_employee_faces(folder_path):
    employee_faces = []
    employee_names = []
    
    for image in os.listdir(folder_path):
        image_path = os.path.join(folder_path, image)
        employee_face = face_recognition.face_encodings(face_recognition.load_image_file(image_path))[0]
        employee_faces.append(employee_face)
        employee_names.append(os.path.splitext(image)[0])
        
    return employee_faces, employee_names

# Function to mark attendance
def mark_attendance(employee_faces, employee_names):
    attendance_file = "C:/Projects/Attendance system/employee_attendance.xlsx"  # Specify full file path
    
    attendance_record = {name: False for name in employee_names}  # Track attendance for each employee
    detected_in_frame = {name: False for name in employee_names}  # Track whether each employee is detected in the current frame
    
    # OpenCV capture from external camera
    video_capture = cv2.VideoCapture(1)
    
    while True:
        ret, frame = video_capture.read()   
        if not ret:
            print("Error: Failed to capture frame")
            break
        
        try:
            face_locations = face_recognition.face_locations(frame)
            face_encodings = face_recognition.face_encodings(frame, face_locations) 

            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):      
                match = face_recognition.compare_faces(employee_faces, face_encoding)
                
                name = None  # Initialize name variable
                
                if True in match:
                    name = employee_names[match.index(True)]
                    
                    if not detected_in_frame[name]:  # Check if person is detected in the frame for the first time
                        timesheet = get_timesheet(name, attendance_file)
                        
                        now = datetime.now()
                        time_string = now.strftime("%H:%M:%S")
                        date_string = now.strftime("%d/%m/%Y")
                        
                        append_entry(timesheet, name, date_string, time_string, attendance_file)

                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                        cv2.putText(frame, f"{name} marked at: {time_string}", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255,255,255), 1)
                        
                        detected_in_frame[name] = True  # Update flag to indicate person is detected in the frame
                    else:
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                        cv2.putText(frame, f"Attendance marked for {name}", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)  # Display different message for already marked attendance
                else:
                    detected_in_frame = {name: False for name in employee_names}  # Reset the entire dictionary if person is not detected in the frame
        except Exception as e:
            print(f"Error: {e}")

        cv2.imshow('Video', frame) 
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the camera capture
    video_capture.release()
    cv2.destroyAllWindows()

# Function to get or create timesheet
def get_timesheet(name, file):
    if not os.path.exists(file):
        wb = openpyxl.Workbook()
        wb.save(file)
    wb = openpyxl.load_workbook(file)
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]

# Function to append attendance entry
def append_entry(ws, name, date, time, file):
    ws.append([name, date, time])
    ws.parent.save(file)

# Main function to run the Streamlit app
def main():
    st.title("Employee Attendance System")

    # Load employee faces
    folder = st.text_input("Enter folder path for employee photos:")
    if folder:
        employee_faces, employee_names = load_employee_faces(folder)

        # Mark attendance
        mark_attendance(employee_faces, employee_names)

if __name__ == "__main__":
    main()
